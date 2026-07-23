from django.test import TestCase
from django.urls import reverse
from django.core import mail
from django.core.exceptions import ValidationError
from django.contrib.admin.sites import AdminSite
from django.contrib.auth.models import User
from io import BytesIO
from unittest.mock import patch
import csv
import datetime
import re
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from rest_framework import status
from rest_framework.test import APITestCase

from datetime import timedelta

from django.utils import timezone as tz

from .models import Event, Lead, _validate_http_https_url
from .admin import LeadAdmin
from .views import (
    _build_google_calendar_url,
    _build_microsoft_calendar_url,
    _build_ics_content,
    _escape_ics_text,
)


class EventModelTestCase(TestCase):
    def setUp(self):
        self.event = Event.objects.create(
            title="Conferencia de Pruebas",
            slug="test-conference",
            notify_email="organizer@example.com",
            name_active=True,
            name_required=True,
            company_active=False,
            company_required=False,
        )

    def test_event_creation(self):
        self.assertEqual(self.event.title, "Conferencia de Pruebas")
        self.assertEqual(self.event.slug, "test-conference")
        self.assertEqual(str(self.event), "Conferencia de Pruebas")


class LeadSubmitAPITestCase(APITestCase):
    def setUp(self):
        self.event = Event.objects.create(
            title="Conferencia de Tecnología",
            slug="tech-conf-2026",
            notify_email="admin@example.com",
            # Fields configuration
            name_active=True,
            name_required=True,
            position_active=True,
            position_required=False,
            email_active=True,
            email_required=True,
            phone_active=False,
            phone_required=False,
            company_active=False,
            company_required=False,
        )
        self.submit_url = reverse("lead-submit", kwargs={"slug": self.event.slug})

    def test_successful_lead_submission(self):
        data = {
            "name": "Juan Perez",
            "email": "juan@example.com",
            "job_position": "Desarrollador",
            "phone": "555-1234",  # Inactive field, will be ignored
            "website": "",  # Empty honeypot
            "terms": True,
        }

        response = self.client.post(self.submit_url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["status"], "ok")

        # Verify Lead creation in database
        lead = Lead.objects.get(event=self.event)
        self.assertEqual(lead.name, "Juan Perez")
        self.assertEqual(lead.email, "juan@example.com")
        self.assertEqual(lead.job_position, "Desarrollador")
        self.assertIsNone(lead.phone)  # Inactive field must be set to None
        self.assertFalse(lead.is_spam)

        # Verify emails dispatched
        self.assertEqual(len(mail.outbox), 2)  # One for admin, one for client
        self.assertEqual(mail.outbox[0].subject, f"Nuevo registro: {self.event.title}")
        self.assertEqual(mail.outbox[1].subject, f"Confirmación de registro: {self.event.title}")

    def test_missing_required_field_returns_400(self):
        data = {
            "name": "",  # Required name is empty
            "email": "juan@example.com",
            "website": "",
            "terms": True,
        }

        response = self.client.post(self.submit_url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("name", response.data["errors"])
        self.assertEqual(
            response.data["errors"]["name"][0],
            "El campo nombre es requerido.",
        )

        # Ensure no lead was saved and no email sent
        self.assertFalse(Lead.objects.filter(event=self.event).exists())
        self.assertEqual(len(mail.outbox), 0)

    def test_honeypot_silently_flags_spam(self):
        data = {
            "name": "Spam Bot",
            "email": "bot@spam.com",
            "website": "http://spam-link.com",  # Honeypot filled!
            "terms": True,
        }

        response = self.client.post(self.submit_url, data, format="json")
        # Returns simulated success (201 Created) to trick the bot
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["status"], "ok")

        # Verify Lead is created in DB but marked as spam
        lead = Lead.objects.get(event=self.event)
        self.assertEqual(lead.name, "Spam Bot")
        self.assertTrue(lead.is_spam)
        # Verify terms is NOT persisted
        self.assertFalse(hasattr(lead, "terms"))

        # Verify NO emails were sent
        self.assertEqual(len(mail.outbox), 0)

    @patch("django.core.mail.EmailMultiAlternatives.send")
    def test_database_save_succeeds_on_smtp_error(self, mock_send):
        # Mock SMTP connection failure
        mock_send.side_effect = Exception("SMTP Connection Refused")

        data = {
            "name": "Maria Lopez",
            "email": "maria@example.com",
            "website": "",
            "terms": True,
        }

        response = self.client.post(self.submit_url, data, format="json")
        # API must return success status, preserving database records
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        # Verify data is still stored safely
        self.assertTrue(Lead.objects.filter(event=self.event, name="Maria Lopez").exists())

    def test_terms_missing_returns_400(self):
        data = {
            "name": "Juan Perez",
            "email": "juan@example.com",
            "website": "",
        }

        response = self.client.post(self.submit_url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("terms", response.data["errors"])
        self.assertIn("Debe aceptar los Términos", str(response.data["errors"]["terms"]))
        self.assertFalse(Lead.objects.filter(event=self.event).exists())

    def test_terms_false_returns_400(self):
        data = {
            "name": "Juan Perez",
            "email": "juan@example.com",
            "website": "",
            "terms": False,
        }

        response = self.client.post(self.submit_url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("terms", response.data["errors"])
        self.assertFalse(Lead.objects.filter(event=self.event).exists())

    def test_terms_not_stored_on_lead_row(self):
        data = {
            "name": "Maria Terms",
            "email": "terms@example.com",
            "website": "",
            "terms": True,
        }

        response = self.client.post(self.submit_url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        lead = Lead.objects.get(email="terms@example.com")
        self.assertFalse(hasattr(lead, "terms"))
        url = reverse("events:event-form", kwargs={"slug": self.event.slug})
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Verify context contains branding settings
        self.assertIn("branding", response.context)
        self.assertEqual(response.context["branding"]["colors"]["primary"], "#0a3c58")

        # Verify template HTML rendering contains brand assets and colors
        content = response.content.decode("utf-8")
        self.assertIn("#0a3c58", content)
        self.assertIn("#072a3e", content)
        self.assertIn("https://aft-reports-generator.s3.amazonaws.com/static/core/imgs/logo-leadforward.jpg", content)

    def test_confirmation_email_branding_and_signature(self):
        data = {
            "name": "Alex Smith",
            "email": "alex@example.com",
            "website": "",
            "terms": True,
        }
        response = self.client.post(self.submit_url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        # Verify client email formatting and signature
        self.assertEqual(len(mail.outbox), 2)
        client_mail = mail.outbox[1]
        self.assertEqual(client_mail.subject, f"Confirmación de registro: {self.event.title}")

        # Text version checks
        self.assertIn("El equipo LeadForward Global Solutions MJ", client_mail.body)

        # HTML version checks
        html_content = client_mail.alternatives[0][0]
        self.assertIn("#0a3c58", html_content)
        self.assertIn("El equipo LeadForward Global Solutions MJ", html_content)
        self.assertIn("https://aft-reports-generator.s3.amazonaws.com/static/core/imgs/logo-leadforward.jpg", html_content)


class InvitationLinkEmailTestCase(APITestCase):
    """Tests for the invitation link CTA in the client confirmation email (now gated through access page)."""

    def setUp(self):
        self.event = Event.objects.create(
            title="Evento con Invitación",
            slug="evento-invitacion",
            notify_email="organizer@example.com",
            name_active=True,
            name_required=True,
            position_active=False,
            position_required=False,
            email_active=True,
            email_required=True,
            phone_active=False,
            phone_required=False,
            event_datetime=tz.now() + timedelta(days=1),
            duration_minutes=60,
        )
        self.submit_url = reverse("lead-submit", kwargs={"slug": self.event.slug})
        self.access_path = reverse("events:event-access", kwargs={"slug": self.event.slug})

    def _post_lead(self, data=None):
        payload = {"name": "Visitante", "email": "v@example.com", "website": "", "terms": True}
        if data:
            payload.update(data)
        return self.client.post(self.submit_url, payload, format="json")

    def test_custom_label_renders_in_client_email(self):
        self.event.invitation_link = "https://zoom.us/j/123?pwd=abc"
        self.event.invitation_label = "Ver grabación"
        self.event.save()

        response = self._post_lead()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        self.assertEqual(len(mail.outbox), 2)
        client_html = mail.outbox[1].alternatives[0][0]
        admin_html = mail.outbox[0].alternatives[0][0]

        self.assertIn("Ver grabación", client_html)
        self.assertIn(self.access_path, client_html)
        self.assertNotIn("https://zoom.us/j/123?pwd=abc", client_html)
        self.assertNotIn("word-break: break-all", client_html)
        self.assertNotIn("Acceder al evento", client_html)
        self.assertNotIn("Ver grabación", admin_html)
        self.assertNotIn(self.access_path, admin_html)

    def test_default_label_used_when_label_blank(self):
        self.event.invitation_link = "https://zoom.us/j/123"
        self.event.invitation_label = ""
        self.event.save()

        response = self._post_lead()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        client_html = mail.outbox[1].alternatives[0][0]
        self.assertIn("Acceder al evento", client_html)
        self.assertIn(self.access_path, client_html)
        self.assertNotIn("https://zoom.us/j/123", client_html)

    def test_no_invitation_link_omits_cta_in_client_email(self):
        self.event.invitation_link = ""
        self.event.save()

        response = self._post_lead()
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        client_html = mail.outbox[1].alternatives[0][0]
        self.assertNotIn("Acceder al evento", client_html)
        self.assertNotIn(self.access_path, client_html)


class InvitationLinkFormTestCase(TestCase):
    """Tests for the invitation link CTA in the public form post-submit success state (now links to access page)."""

    def setUp(self):
        self.event = Event.objects.create(
            title="Evento Iframe",
            slug="evento-iframe",
            notify_email="organizer@example.com",
            name_active=True,
            name_required=True,
            event_datetime=tz.now() + timedelta(days=1),
            duration_minutes=60,
        )
        self.form_url = reverse("events:event-form", kwargs={"slug": self.event.slug})
        self.access_path = reverse("events:event-access", kwargs={"slug": self.event.slug})

    def test_form_renders_invitation_cta_when_link_and_datetime_set(self):
        self.event.invitation_link = "https://zoom.us/j/123"
        self.event.invitation_label = "Unirse ahora"
        self.event.save()

        response = self.client.get(self.form_url)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode("utf-8")
        self.assertIn('id="invitation-cta"', content)
        self.assertIn('id="invitation-cta-link"', content)
        self.assertIn(self.access_path, content)
        self.assertIn("Unirse ahora", content)
        self.assertNotIn('ctaLink.href = "https://zoom.us/j/123"', content)
        self.assertIn('id="invitation-cta" class="alert alert-success" style="display: none;', content)

    def test_form_omits_invitation_cta_when_no_link(self):
        self.event.invitation_link = ""
        self.event.save()

        response = self.client.get(self.form_url)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode("utf-8")
        self.assertNotIn('id="invitation-cta"', content)
        self.assertNotIn("Acceder al evento", content)

    def test_form_omits_invitation_cta_when_no_datetime(self):
        self.event.event_datetime = None
        self.event.save()

        response = self.client.get(self.form_url)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode("utf-8")
        self.assertNotIn('id="invitation-cta"', content)

    def test_form_uses_default_label_when_label_blank(self):
        self.event.invitation_link = "https://zoom.us/j/123"
        self.event.invitation_label = ""
        self.event.save()

        response = self.client.get(self.form_url)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode("utf-8")
        self.assertIn("ctaLink.textContent = \"Acceder al evento\"", content)


class InvitationLinkValidationTestCase(TestCase):
    """Tests for the http(s)-only URL validator on Event.invitation_link."""

    def test_javascript_scheme_rejected(self):
        e = Event(
            title="x", slug="x-js", notify_email="a@b.com",
            invitation_link="javascript:alert(1)",
        )
        with self.assertRaises(ValidationError) as ctx:
            e.full_clean()
        # Spec scenario 7: error message mentions http(s) validity
        self.assertIn("http", str(ctx.exception))

    def test_invalid_url_rejected(self):
        e = Event(
            title="x", slug="x-bad", notify_email="a@b.com",
            invitation_link="not-a-url",
        )
        with self.assertRaises(ValidationError) as ctx:
            e.full_clean()
        self.assertIn("http", str(ctx.exception))

    def test_ftp_scheme_rejected(self):
        e = Event(
            title="x", slug="x-ftp", notify_email="a@b.com",
            invitation_link="ftp://files.example.com/y",
        )
        with self.assertRaises(ValidationError) as ctx:
            e.full_clean()
        self.assertIn("http", str(ctx.exception))

    def test_https_url_with_query_and_utm_accepted(self):
        e = Event.objects.create(
            title="x", slug="x-ok", notify_email="a@b.com",
            invitation_link="https://zoom.us/j/123?pwd=abc&utm_source=mail",
        )
        self.assertEqual(
            e.invitation_link,
            "https://zoom.us/j/123?pwd=abc&utm_source=mail",
        )

    def test_long_url_within_500_chars_accepted(self):
        long_path = "a" * 200
        url = f"https://example.com/{long_path}?utm_source=mail&utm_campaign=q3"
        self.assertLessEqual(len(url), 500)
        e = Event.objects.create(
            title="x", slug="x-long", notify_email="a@b.com",
            invitation_link=url,
        )
        self.assertEqual(e.invitation_link, url)

    def test_empty_invitation_link_allowed(self):
        e = Event.objects.create(
            title="x", slug="x-empty", notify_email="a@b.com",
            invitation_link="", invitation_label="",
        )
        self.assertFalse(e.invitation_link)
        self.assertEqual(e.invitation_label, "")


class InvitationLinkAdminTestCase(TestCase):
    """Tests for admin edit form and list view exposing the new fields."""

    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="admin"
        )
        self.client.login(username="admin", password="admin")

    def test_admin_edit_form_exposes_all_fields(self):
        event = Event.objects.create(
            title="Admin Test", slug="admin-test",
            notify_email="o@e.com",
        )
        url = reverse("admin:events_event_change", args=[event.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode("utf-8")
        self.assertIn('name="invitation_link"', content)
        self.assertIn('name="invitation_label"', content)
        self.assertIn('name="event_datetime_0"', content)
        self.assertIn('name="event_datetime_1"', content)
        self.assertIn('name="duration_minutes"', content)
        self.assertIn("Enlace de invitación", content)
        self.assertIn("Texto del botón de invitación", content)
        self.assertIn("Fecha y hora del evento", content)
        self.assertIn("Duración (minutos)", content)

    def test_admin_list_view_renders_clickable_link(self):
        Event.objects.create(
            title="With Link", slug="with-link",
            notify_email="o@e.com",
            invitation_link="https://zoom.us/j/123",
        )
        Event.objects.create(
            title="No Link", slug="no-link",
            notify_email="o@e.com",
        )
        url = reverse("admin:events_event_changelist")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode("utf-8")
        # Clickable link with target/rel
        self.assertIn('href="https://zoom.us/j/123"', content)
        self.assertIn('target="_blank"', content)
        self.assertIn('rel="noopener noreferrer"', content)
        # Em-dash placeholder for the no-link event
        self.assertIn("—", content)

    def test_admin_post_persists_invitation_link_and_label(self):
        """Spec scenario 12: saving the form persists the new values."""
        event = Event.objects.create(
            title="Post Test", slug="post-test",
            notify_email="o@e.com",
            name_active=True, name_required=True,
            position_active=True, position_required=True,
            email_active=True, email_required=True,
            phone_active=True, phone_required=True,
            company_active=False, company_required=False,
        )
        url = reverse("admin:events_event_change", args=[event.pk])
        response = self.client.post(url, {
            "title": "Post Test Updated",
            "slug": "post-test",
            "is_active": "on",
            "notify_email": "o@e.com",
            "invitation_link": "https://zoom.us/j/999",
            "invitation_label": "Únete ya",
            "event_datetime_0": "",
            "event_datetime_1": "",
            "duration_minutes": "0",
            # The remaining fieldsets are in the admin form; send their
            # current values so the form does not blank them out.
            "name_active": "on", "name_required": "on",
            "position_active": "on", "position_required": "on",
            "email_active": "on", "email_required": "on",
            "phone_active": "on", "phone_required": "on",
            "_save": "Save",
        })
        # Redirect after successful save
        self.assertIn(response.status_code, (302, 200))
        event.refresh_from_db()
        self.assertEqual(event.invitation_link, "https://zoom.us/j/999")
        self.assertEqual(event.invitation_label, "Únete ya")


class InvitationLabelDisplayPropertyTestCase(TestCase):
    """Tests for the Event.invitation_label_display property (W1 fix)."""

    def test_empty_label_returns_default(self):
        e = Event.objects.create(
            title="x", slug="x-empty", notify_email="a@b.com",
            invitation_label="",
        )
        self.assertEqual(e.invitation_label_display, "Acceder al evento")

    def test_whitespace_only_label_returns_default(self):
        e = Event.objects.create(
            title="x", slug="x-ws", notify_email="a@b.com",
            invitation_label="   ",
        )
        self.assertEqual(e.invitation_label_display, "Acceder al evento")

    def test_none_label_returns_default(self):
        # Model uses default="" (NOT null), so we test the property with
        # an in-memory instance rather than a saved DB row.
        e = Event(invitation_label=None)
        self.assertEqual(e.invitation_label_display, "Acceder al evento")

    def test_label_with_surrounding_whitespace_is_trimmed(self):
        e = Event.objects.create(
            title="x", slug="x-trim", notify_email="a@b.com",
            invitation_label="  Ver grabación  ",
        )
        self.assertEqual(e.invitation_label_display, "Ver grabación")

    def test_label_without_whitespace_returned_as_is(self):
        e = Event.objects.create(
            title="x", slug="x-pure", notify_email="a@b.com",
            invitation_label="Ver grabación",
        )
        self.assertEqual(e.invitation_label_display, "Ver grabación")


class InvitationLabelEscapejsSafetyTestCase(TestCase):
    """W2 fix: malicious labels must not break the inline <script> block."""

    def setUp(self):
        self.event = Event.objects.create(
            title="Escape Test", slug="escape-test",
            notify_email="a@b.com", name_active=True, name_required=True,
            event_datetime=tz.now() + timedelta(days=1),
            duration_minutes=60,
        )
        self.form_url = reverse("events:event-form", kwargs={"slug": self.event.slug})

    def test_invitation_label_escapejs_safe(self):
        evil = "</script><script>alert(1)</script>"
        self.event.invitation_link = "https://example.com"
        self.event.invitation_label = evil
        self.event.save()

        response = self.client.get(self.form_url)
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")

        self.assertIn(r"\u003C", content)
        self.assertNotIn("alert(1)<", content)


class LongInvitationLinkRenderTestCase(APITestCase):
    """S2 fix: long URL (up to 500 chars) must not leak into email or form surfaces."""

    def setUp(self):
        path = "p" * 400
        self.long_url = f"https://example.com/{path}?utm_source=mail&utm_campaign=q3"
        self.assertLessEqual(len(self.long_url), 500)
        self.assertGreater(len(self.long_url), 400)

        self.event = Event.objects.create(
            title="Long URL Event", slug="long-url-event",
            notify_email="o@e.com",
            name_active=True, name_required=True,
            position_active=False, position_required=False,
            phone_active=False, phone_required=False,
            email_active=True, email_required=True,
            invitation_link=self.long_url,
            event_datetime=tz.now() + timedelta(days=1),
            duration_minutes=60,
        )
        self.submit_url = reverse("lead-submit", kwargs={"slug": self.event.slug})

    def test_long_url_not_exposed_in_client_email(self):
        response = self.client.post(self.submit_url, {
            "name": "L", "email": "l@e.com", "website": "", "terms": True,
        }, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        client_html = mail.outbox[1].alternatives[0][0]
        self.assertNotIn("example.com/p", client_html)
        self.assertNotIn("utm_source", client_html)

    def test_access_url_rendered_in_iframe_template(self):
        form_url = reverse("events:event-form", kwargs={"slug": self.event.slug})
        access_path = reverse("events:event-access", kwargs={"slug": self.event.slug})
        response = self.client.get(form_url)
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn(access_path, content)
        self.assertNotIn(self.long_url.split("?")[0], content)


class LeadExportActionsTestCase(TestCase):
    """Tests for the LeadAdmin CSV/Excel export actions.

    Covers:
    - Task 1.4 / 4.2: CSV byte-level regression lock.
    - Tasks 3.x: Excel action header, content, MIME, sheet title, bold, empty.
    - Task 4.1: CSV Content-Type / Content-Disposition preservation.
    - Tasks 5.1 / 5.2: dropdown listing & no-rows-selected behavior.
    """

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="admin"
        )
        cls.event = Event.objects.create(
            title="Conferencia Exportable",
            slug="export-conf",
            notify_email="organizer@example.com",
        )

    def setUp(self):
        self.lead_non_spam = Lead.objects.create(
            event=self.event,
            name="Juan Perez",
            email="juan@example.com",
            phone="555-1234",
            job_position="Desarrollador",
            company="Empresa S.A.",
            is_spam=False,
        )
        self.lead_spam = Lead.objects.create(
            event=self.event,
            name="Spam Bot",
            email="bot@spam.com",
            phone="555-0000",
            job_position="Bot",
            company="Spam Inc.",
            is_spam=True,
        )
        # Freeze created_at deterministically so strftime-based regression
        # locks aren't at the mercy of auto_now_add's wall-clock value.
        from django.utils import timezone
        fixed = timezone.make_aware(
            __import__("datetime").datetime(2026, 7, 20, 14, 32, 5)
        )
        Lead.objects.filter(pk=self.lead_non_spam.pk).update(created_at=fixed)
        Lead.objects.filter(pk=self.lead_spam.pk).update(
            created_at=fixed + __import__("datetime").timedelta(hours=1)
        )
        self.lead_non_spam.refresh_from_db()
        self.lead_spam.refresh_from_db()

        self.site = AdminSite()
        self.modeladmin = LeadAdmin(Lead, self.site)
        # Build a representative queryset: spam first, then non-spam, mirroring
        # the iteration order produced by the admin's "select all in filter".
        self.queryset = Lead.objects.order_by("-is_spam", "id")

    # ----- CSV regression baseline (tasks 1.4, 4.1, 4.2) -----

    def _expected_csv_bytes(self):
        """Build the exact CSV bytes the action is expected to produce.

        Mirrors the original inline implementation byte-for-byte so the
        refactored action must match it. Notably Django's
        ``HttpResponse(content_type="text/csv; charset=utf-8-sig")`` encodes
        each ``write()`` call with the UTF-8-sig codec, which prepends the
        BOM before every row — not only the first one. Reproducing that
        requires writing through the same HttpResponse pipeline.
        """
        from django.http import HttpResponse
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        writer = csv.writer(response)
        writer.writerow([
            "Nombre",
            "Email",
            "Teléfono",
            "Puesto de trabajo",
            "Empresa",
            "Evento",
            "Spam",
            "Fecha de Registro",
        ])
        for obj in self.queryset:
            writer.writerow([
                obj.name or "",
                obj.email or "",
                obj.phone or "",
                obj.job_position or "",
                obj.company or "",
                obj.event.title,
                "Sí" if obj.is_spam else "No",
                obj.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ])
        return response.content

    def test_csv_action_registered_on_actions(self):
        self.assertIn("export_as_csv", LeadAdmin.actions)
        self.assertIn("export_as_excel", LeadAdmin.actions)

    def test_csv_action_headers_and_bytes_unchanged(self):
        request = None  # action doesn't read request
        response = self.modeladmin.export_as_csv(request, self.queryset)

        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8-sig")
        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="leads_registro.csv"',
        )
        self.assertEqual(response.content, self._expected_csv_bytes())

    # ----- Excel action tests (tasks 3.1–3.6) -----

    def _load_xlsx(self, response):
        return load_workbook(BytesIO(response.content))

    def test_excel_action_response_headers(self):
        request = None
        response = self.modeladmin.export_as_excel(request, self.queryset)

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="leads_registro.xlsx"',
        )

    def test_excel_workbook_single_sheet_named_after_verbose_plural(self):
        request = None
        response = self.modeladmin.export_as_excel(request, self.queryset)

        wb = self._load_xlsx(response)
        self.assertEqual(len(wb.worksheets), 1)
        self.assertEqual(wb.active.title, Lead._meta.verbose_name_plural)
        self.assertEqual(wb.active.title, "Registros (Leads)")

    def test_excel_header_row_columns_bold_and_order(self):
        request = None
        response = self.modeladmin.export_as_excel(request, self.queryset)

        sheet = self._load_xlsx(response).active
        headers = [sheet.cell(row=1, column=c).value for c in range(1, 9)]
        self.assertEqual(headers, LeadAdmin._HEADER_COLUMNS)
        for c in range(1, 9):
            self.assertTrue(
                sheet.cell(row=1, column=c).font.bold,
                f"Header cell column {c} should be bold",
            )

    def test_excel_data_rows_match_lead_rows(self):
        request = None
        response = self.modeladmin.export_as_excel(request, self.queryset)

        sheet = self._load_xlsx(response).active
        # Two data rows for our two leads + 1 header row
        self.assertEqual(sheet.max_row, 3)
        # Iterate queryset in the same order the admin action did
        for i, obj in enumerate(self.queryset, start=2):
            expected = [
                obj.name or "",
                obj.email or "",
                obj.phone or "",
                obj.job_position or "",
                obj.company or "",
                obj.event.title,
                "Sí" if obj.is_spam else "No",
                obj.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
            actual = [sheet.cell(row=i, column=c).value for c in range(1, 9)]
            self.assertEqual(actual, expected)

    def test_excel_spam_column_renders_spanish_strings(self):
        request = None
        response = self.modeladmin.export_as_excel(request, self.queryset)
        sheet = self._load_xlsx(response).active

        spam_values = {sheet.cell(row=r, column=7).value for r in range(2, sheet.max_row + 1)}
        self.assertEqual(spam_values, {"Sí", "No"})

    def test_excel_date_column_renders_as_formatted_string(self):
        request = None
        response = self.modeladmin.export_as_excel(request, self.queryset)
        sheet = self._load_xlsx(response).active

        for r in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r, column=8).value
            self.assertIsInstance(val, str)
            # Match the exact format pattern
            self.assertRegex(val, r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$")

    def test_excel_empty_queryset_returns_only_header_row(self):
        request = None
        empty_qs = Lead.objects.none()
        response = self.modeladmin.export_as_excel(request, empty_qs)

        sheet = self._load_xlsx(response).active
        self.assertEqual(sheet.max_row, 1)
        headers = [sheet.cell(row=1, column=c).value for c in range(1, 9)]
        self.assertEqual(headers, LeadAdmin._HEADER_COLUMNS)
        for c in range(1, 9):
            self.assertTrue(sheet.cell(row=1, column=c).font.bold)
        # HTTP semantics still correct
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ----- Cross-action dropdown + no-rows (tasks 5.1, 5.2) -----

    def test_changelist_lists_both_export_actions(self):
        self.client.login(username="admin", password="admin")
        url = reverse("admin:events_lead_changelist")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode("utf-8")
        self.assertIn("Exportar registros seleccionados a CSV", content)
        self.assertIn("Exportar registros seleccionados a Excel", content)

    def test_export_with_no_rows_selected_shows_no_items_message(self):
        self.client.login(username="admin", password="admin")
        url = reverse("admin:events_lead_changelist")
        # POST the action without _selected_action rows
        response = self.client.post(url, {
            "action": "export_as_excel",
            "_selected_action": [],
            "post": "yes",
            "select_across": "0",
            "index": "0",
        })
        # Django redirects back to changelist with the no-items-selected msg
        self.assertEqual(response.status_code, 302)

        # Follow the redirect and confirm the standard admin message appears.
        # Jazzmin translates the message to Spanish ("Deben existir items
        # seleccionados para poder realizar acciones sobre los mismos. No
        # se modificó ningún item."), so accept either language's wording.
        follow = self.client.get(response["Location"])
        body = follow.content.decode("utf-8")
        self.assertTrue(
            "No items selected" in body or "items seleccionados" in body,
            "Expected no-items-selected admin warning, got body snippet: "
                + body[:500],
        )


class EventAccessGateTestCase(TestCase):
    """Tests for the EventAccessView intermediate access page."""

    def setUp(self):
        self.access_url = reverse("events:event-access", kwargs={"slug": "test-access"})

    def _event(self, **overrides):
        params = dict(
            title="Evento de Acceso",
            slug="test-access",
            notify_email="o@e.com",
            invitation_link="https://zoom.us/j/123",
            event_datetime=tz.now() + timedelta(hours=3),
            duration_minutes=60,
        )
        params.update(overrides)
        return Event.objects.create(**params)

    def test_redirects_within_one_hour(self):
        self._event(event_datetime=tz.now() + timedelta(minutes=30))
        response = self.client.get(self.access_url)
        self.assertRedirects(response, "https://zoom.us/j/123", fetch_redirect_response=False)

    def test_renders_countdown_when_more_than_one_hour(self):
        self._event(event_datetime=tz.now() + timedelta(hours=3))
        response = self.client.get(self.access_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "events/access.html")
        self.assertContains(response, "Evento de Acceso")
        self.assertContains(response, "invitation-btn")
        self.assertContains(response, "countdown")

    def test_renders_ended_when_event_passed_with_duration(self):
        self._event(
            event_datetime=tz.now() - timedelta(hours=2),
            duration_minutes=60,
        )
        response = self.client.get(self.access_url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "events/ended.html")
        self.assertContains(response, "ya ha finalizado")

    def test_skips_ended_check_when_duration_is_zero(self):
        self._event(
            event_datetime=tz.now() - timedelta(hours=2),
            duration_minutes=0,
        )
        response = self.client.get(self.access_url)
        self.assertRedirects(response, "https://zoom.us/j/123", fetch_redirect_response=False)

    def test_404_when_invitation_link_missing(self):
        self._event(invitation_link="")
        response = self.client.get(self.access_url)
        self.assertEqual(response.status_code, 404)

    def test_404_when_event_datetime_missing(self):
        self._event(event_datetime=None)
        response = self.client.get(self.access_url)
        self.assertEqual(response.status_code, 404)


class EventModelValidationTestCase(TestCase):
    """Tests for Event model clean() validation."""

    def test_rejects_duration_zero_when_datetime_set(self):
        e = Event(
            title="x", slug="x-dur", notify_email="a@b.com",
            event_datetime=tz.now() + timedelta(days=1),
            duration_minutes=0,
        )
        with self.assertRaises(ValidationError):
            e.clean()

    def test_allows_duration_zero_when_datetime_null(self):
        e = Event(
            title="x", slug="x-null", notify_email="a@b.com",
            event_datetime=None,
            duration_minutes=0,
        )
        try:
            e.clean()
        except ValidationError:
            self.fail("clean() raised ValidationError when event_datetime is None")


class CalendarUrlHelpersTestCase(TestCase):
    def setUp(self):
        self.event = Event.objects.create(
            title="Conferencia de Pruebas",
            slug="test-calendar",
            notify_email="organizer@example.com",
            invitation_link="https://zoom.us/j/123",
            event_datetime=tz.now() + timedelta(days=7),
            duration_minutes=120,
        )

    def test_google_url_contains_required_params(self):
        url = _build_google_calendar_url(self.event)
        self.assertIn("action=TEMPLATE", url)
        self.assertIn("text=", url)
        self.assertIn("ctz=America%2FMexico_City", url)
        self.assertIn("trp=false", url)
        self.assertIn("location=https%3A%2F%2Fzoom.us%2Fj%2F123", url)

    def test_google_url_omits_location_when_no_link(self):
        self.event.invitation_link = None
        self.event.save()
        url = _build_google_calendar_url(self.event)
        self.assertNotIn("location=", url)

    def test_microsoft_url_contains_required_params(self):
        url = _build_microsoft_calendar_url(self.event)
        self.assertIn("subject=", url)
        self.assertIn("startdt=", url)
        self.assertIn("enddt=", url)
        self.assertIn("path=%2Fcalendar%2Faction%2Fcompose", url)
        self.assertIn("rru=addevent", url)
        self.assertIn("location=https%3A%2F%2Fzoom.us%2Fj%2F123", url)

    def test_microsoft_url_omits_location_when_no_link(self):
        self.event.invitation_link = None
        self.event.save()
        url = _build_microsoft_calendar_url(self.event)
        self.assertNotIn("location=", url)

    def test_ics_content_contains_required_fields(self):
        content = _build_ics_content(self.event)
        self.assertIn("BEGIN:VCALENDAR", content)
        self.assertIn("DTSTART:", content)
        self.assertIn("DTEND:", content)
        self.assertIn("SUMMARY:Conferencia de Pruebas", content)
        self.assertIn("UID:test-calendar@aft-dashboard", content)
        self.assertIn("LOCATION:https://zoom.us/j/123", content)
        self.assertIn("END:VCALENDAR", content)
        self.assertTrue(content.endswith("\r\n"))

    def test_ics_content_omits_location_when_no_link(self):
        self.event.invitation_link = None
        self.event.save()
        content = _build_ics_content(self.event)
        self.assertNotIn("LOCATION:", content)

    def test_ics_content_dtend_equals_dtstart_when_duration_zero(self):
        self.event.duration_minutes = 0
        self.event.save()
        content = _build_ics_content(self.event)
        lines = content.split("\r\n")
        dtstart = [l for l in lines if l.startswith("DTSTART:")][0]
        dtend = [l for l in lines if l.startswith("DTEND:")][0]
        self.assertEqual(dtstart.split(":")[1], dtend.split(":")[1])

    def test_ics_escapes_special_characters(self):
        self.event.title = "Evento; Especial, Edicion\\2024"
        self.event.save()
        content = _build_ics_content(self.event)
        self.assertIn("SUMMARY:Evento\\; Especial\\, Edicion\\\\2024", content)

    def test_ics_escapes_location_special_characters(self):
        self.event.invitation_link = "https://zoom.us/j/123;extra,param\\back"
        self.event.save()
        content = _build_ics_content(self.event)
        self.assertIn("LOCATION:https://zoom.us/j/123\\;extra\\,param\\\\back", content)

    def test_calendar_urls_contain_correct_utc_dates(self):
        mexico_tz = ZoneInfo("America/Mexico_City")
        dt = datetime.datetime(2024, 3, 15, 16, 0, 0, tzinfo=mexico_tz)
        self.event.event_datetime = dt
        self.event.duration_minutes = 60
        self.event.save()

        google_url = _build_google_calendar_url(self.event)
        self.assertIn("dates=20240315T220000Z%2F20240315T230000Z", google_url)

        microsoft_url = _build_microsoft_calendar_url(self.event)
        self.assertIn("startdt=2024-03-15T22%3A00%3A00%2B00%3A00", microsoft_url)
        self.assertIn("enddt=2024-03-15T23%3A00%3A00%2B00%3A00", microsoft_url)

        ics_content = _build_ics_content(self.event)
        self.assertIn("DTSTART:20240315T220000Z", ics_content)
        self.assertIn("DTEND:20240315T230000Z", ics_content)


class EventCalendarIcsViewTestCase(TestCase):
    def setUp(self):
        self.event = Event.objects.create(
            title="Evento ICS",
            slug="test-ics",
            notify_email="organizer@example.com",
            invitation_link="https://zoom.us/j/123",
            event_datetime=tz.now() + timedelta(days=7),
            duration_minutes=60,
        )
        self.ics_url = reverse("events:event-ics", kwargs={"slug": self.event.slug})

    def test_ics_view_returns_correct_headers(self):
        response = self.client.get(self.ics_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/calendar")
        self.assertIn('attachment; filename="test-ics.ics"', response["Content-Disposition"])

    def test_ics_view_body_contains_icalendar(self):
        response = self.client.get(self.ics_url)
        self.assertContains(response, "BEGIN:VCALENDAR")
        self.assertContains(response, "SUMMARY:Evento ICS")
        self.assertContains(response, "END:VCALENDAR")

    def test_ics_view_404_when_event_inactive(self):
        self.event.is_active = False
        self.event.save()
        response = self.client.get(self.ics_url)
        self.assertEqual(response.status_code, 404)

    def test_ics_view_404_when_datetime_missing(self):
        self.event.event_datetime = None
        self.event.save()
        response = self.client.get(self.ics_url)
        self.assertEqual(response.status_code, 404)


class CalendarButtonsTemplateTestCase(TestCase):
    def setUp(self):
        self.event = Event.objects.create(
            title="Evento Calendario",
            slug="cal-event",
            notify_email="organizer@example.com",
            invitation_link="https://zoom.us/j/123",
            event_datetime=tz.now() + timedelta(hours=3),
            duration_minutes=120,
        )
        self.access_url = reverse("events:event-access", kwargs={"slug": self.event.slug})

    def test_access_page_renders_calendar_buttons(self):
        response = self.client.get(self.access_url)
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("calendar-section", content)
        self.assertIn("Google Calendar", content)
        self.assertIn("Outlook", content)
        self.assertIn("Apple Calendar", content)

        google_btn = re.search(r'<a[^>]*google\.com/calendar[^>]*>.*?Google Calendar.*?</a>', content, re.DOTALL)
        self.assertIsNotNone(google_btn)
        self.assertIn('target="_blank"', google_btn.group())
        self.assertNotIn("download", google_btn.group())

        outlook_btn = re.search(r'<a[^>]*outlook\.office\.com[^>]*>.*?Outlook.*?</a>', content, re.DOTALL)
        self.assertIsNotNone(outlook_btn)
        self.assertIn('target="_blank"', outlook_btn.group())
        self.assertNotIn("download", outlook_btn.group())

        apple_btn = re.search(r'<a[^>]*download[^>]*>.*?Apple Calendar.*?</a>', content, re.DOTALL)
        self.assertIsNotNone(apple_btn)
        self.assertIn("download", apple_btn.group())
        self.assertNotIn('target="_blank"', apple_btn.group())

    def test_access_page_404_when_datetime_missing(self):
        self.event.event_datetime = None
        self.event.save()
        response = self.client.get(self.access_url)
        self.assertEqual(response.status_code, 404)

